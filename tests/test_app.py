"""
民泊管理アプリ 自動テスト (Playwright + Excel出力)

使い方:
  1. pip install -r requirements.txt
  2. playwright install chromium
  3. python test_app.py

設定:
  - APP_URL: テスト対象のアプリURL（GASデプロイURL）
  - EXCEL_OUTPUT: 結果出力先のExcelファイルパス
"""

import sys
import traceback
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# 設定 - テスト対象のURLをここに入力してください
# ============================================================
APP_URL = "https://script.google.com/macros/s/YOUR_DEPLOYMENT_ID/exec"

# Excel出力先
EXCEL_OUTPUT = Path(__file__).parent / "test_results.xlsx"

# タイムアウト（ミリ秒）
PAGE_LOAD_TIMEOUT = 30000
ACTION_TIMEOUT = 10000

# テスト入力値
TEST_INPUT_TEXT = "テスト入力"


# ============================================================
# テスト結果を格納するクラス
# ============================================================
class TestResult:
    def __init__(self, step_name: str):
        self.step_name = step_name
        self.status = "未実行"
        self.detail = ""
        self.timestamp = ""
        self.screenshot_path = ""

    def mark_pass(self, detail: str = ""):
        self.status = "OK"
        self.detail = detail
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def mark_fail(self, detail: str = ""):
        self.status = "NG"
        self.detail = detail
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def mark_skip(self, detail: str = ""):
        self.status = "スキップ"
        self.detail = detail
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ============================================================
# Excel書き出し
# ============================================================
def write_results_to_excel(results: list[TestResult], output_path: Path):
    """テスト結果をExcelファイルに書き出す"""
    wb = Workbook()
    ws = wb.active
    ws.title = "テスト結果"

    # ヘッダースタイル
    header_font = Font(name="Meiryo", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # ヘッダー行
    headers = ["No.", "テストステップ", "結果", "詳細", "実行日時", "スクリーンショット"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 結果スタイル
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ok_font = Font(name="Meiryo", color="006100")
    ng_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ng_font = Font(name="Meiryo", color="9C0006")
    skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    skip_font = Font(name="Meiryo", color="9C6500")
    body_font = Font(name="Meiryo", size=10)

    for i, result in enumerate(results, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = body_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=result.step_name).font = body_font
        status_cell = ws.cell(row=row, column=3, value=result.status)
        ws.cell(row=row, column=4, value=result.detail).font = body_font
        ws.cell(row=row, column=5, value=result.timestamp).font = body_font
        ws.cell(row=row, column=6, value=result.screenshot_path).font = body_font

        # 結果セルに色を付ける
        if result.status == "OK":
            status_cell.fill = ok_fill
            status_cell.font = ok_font
        elif result.status == "NG":
            status_cell.fill = ng_fill
            status_cell.font = ng_font
        else:
            status_cell.fill = skip_fill
            status_cell.font = skip_font
        status_cell.alignment = Alignment(horizontal="center")

        # 罫線
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border

    # 列幅調整
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40

    # サマリー行
    total = len(results)
    ok_count = sum(1 for r in results if r.status == "OK")
    ng_count = sum(1 for r in results if r.status == "NG")
    skip_count = sum(1 for r in results if r.status == "スキップ")
    summary_row = total + 3

    ws.cell(row=summary_row, column=1, value="【結果サマリー】").font = Font(
        name="Meiryo", bold=True, size=11
    )
    ws.cell(row=summary_row + 1, column=1, value=f"  合計: {total} ステップ").font = body_font
    ws.cell(row=summary_row + 2, column=1, value=f"  OK: {ok_count}").font = Font(
        name="Meiryo", color="006100"
    )
    ws.cell(row=summary_row + 3, column=1, value=f"  NG: {ng_count}").font = Font(
        name="Meiryo", color="9C0006"
    )
    ws.cell(row=summary_row + 4, column=1, value=f"  スキップ: {skip_count}").font = Font(
        name="Meiryo", color="9C6500"
    )

    wb.save(output_path)
    print(f"\n結果をExcelに保存しました: {output_path}")


# ============================================================
# スクリーンショット保存
# ============================================================
def take_screenshot(page, name: str) -> str:
    """スクリーンショットを保存してパスを返す"""
    screenshots_dir = Path(__file__).parent / "screenshots"
    screenshots_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = screenshots_dir / f"{name}_{timestamp}.png"
    page.screenshot(path=str(path), full_page=False)
    return str(path)


# ============================================================
# メインテスト処理
# ============================================================
def run_tests():
    results: list[TestResult] = []
    console_errors: list[str] = []

    print("=" * 60)
    print("民泊管理アプリ 自動テスト開始")
    print(f"対象URL: {APP_URL}")
    print(f"実行日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    with sync_playwright() as p:
        # --------------------------------------------------
        # Step 1: ブラウザ起動
        # --------------------------------------------------
        step1 = TestResult("ブラウザ起動")
        results.append(step1)
        try:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                viewport={"width": 1280, "height": 720},
                locale="ja-JP",
            )
            page = context.new_page()

            # コンソールエラーを収集
            page.on("console", lambda msg: (
                console_errors.append(f"[{msg.type}] {msg.text}")
                if msg.type == "error" else None
            ))
            # ページエラー（未処理例外）を収集
            page.on("pageerror", lambda err: console_errors.append(f"[PageError] {err}"))

            step1.mark_pass("Chromium (headless) 起動成功")
            print(f"  [OK] {step1.step_name}")
        except Exception as e:
            step1.mark_fail(f"ブラウザ起動失敗: {e}")
            print(f"  [NG] {step1.step_name}: {e}")
            write_results_to_excel(results, EXCEL_OUTPUT)
            return 1

        try:
            # --------------------------------------------------
            # Step 2: ページ読み込み
            # --------------------------------------------------
            step2 = TestResult("ページ読み込み")
            results.append(step2)
            try:
                page.goto(APP_URL, timeout=PAGE_LOAD_TIMEOUT, wait_until="networkidle")
                step2.mark_pass(f"ページ読み込み完了 (タイトル: {page.title()})")
                step2.screenshot_path = take_screenshot(page, "01_page_loaded")
                print(f"  [OK] {step2.step_name}")
            except PlaywrightTimeout:
                step2.mark_fail(f"タイムアウト ({PAGE_LOAD_TIMEOUT}ms) - ページが読み込めませんでした")
                step2.screenshot_path = take_screenshot(page, "01_page_timeout")
                print(f"  [NG] {step2.step_name}: タイムアウト")
                write_results_to_excel(results, EXCEL_OUTPUT)
                return 1
            except Exception as e:
                step2.mark_fail(f"ページ読み込みエラー: {e}")
                print(f"  [NG] {step2.step_name}: {e}")
                write_results_to_excel(results, EXCEL_OUTPUT)
                return 1

            # --------------------------------------------------
            # Step 3: テキストボックスを探す
            # --------------------------------------------------
            step3 = TestResult("テキストボックス検出")
            results.append(step3)
            text_input = None
            try:
                # 表示されているテキスト入力を探す（優先順位付き）
                selectors = [
                    'input[type="text"]:visible',
                    'textarea:visible',
                    'input:not([type="hidden"]):not([type="password"]):not([type="checkbox"]):not([type="radio"]):visible',
                ]
                for selector in selectors:
                    locator = page.locator(selector).first
                    if locator.count() > 0 and locator.is_visible():
                        text_input = locator
                        break

                if text_input is None:
                    step3.mark_fail("表示されているテキストボックスが見つかりません")
                    step3.screenshot_path = take_screenshot(page, "02_no_textbox")
                    print(f"  [NG] {step3.step_name}")
                else:
                    tag = text_input.evaluate("el => el.tagName.toLowerCase()")
                    input_id = text_input.evaluate("el => el.id || '(id無し)'")
                    placeholder = text_input.evaluate("el => el.placeholder || '(placeholder無し)'")
                    step3.mark_pass(f"検出: <{tag}> id={input_id}, placeholder={placeholder}")
                    print(f"  [OK] {step3.step_name}: <{tag}> id={input_id}")
            except Exception as e:
                step3.mark_fail(f"テキストボックス検出エラー: {e}")
                step3.screenshot_path = take_screenshot(page, "02_textbox_error")
                print(f"  [NG] {step3.step_name}: {e}")

            # --------------------------------------------------
            # Step 4: テキスト入力
            # --------------------------------------------------
            step4 = TestResult(f"テキスト入力 ('{TEST_INPUT_TEXT}')")
            results.append(step4)
            if text_input is not None:
                try:
                    text_input.click()
                    text_input.fill(TEST_INPUT_TEXT)
                    # 入力値を確認
                    actual = text_input.input_value()
                    if actual == TEST_INPUT_TEXT:
                        step4.mark_pass(f"入力成功: '{actual}'")
                        step4.screenshot_path = take_screenshot(page, "03_text_entered")
                        print(f"  [OK] {step4.step_name}")
                    else:
                        step4.mark_fail(f"入力値が一致しません (期待: '{TEST_INPUT_TEXT}', 実際: '{actual}')")
                        step4.screenshot_path = take_screenshot(page, "03_text_mismatch")
                        print(f"  [NG] {step4.step_name}: 値不一致")
                except Exception as e:
                    step4.mark_fail(f"テキスト入力エラー: {e}")
                    step4.screenshot_path = take_screenshot(page, "03_text_error")
                    print(f"  [NG] {step4.step_name}: {e}")
            else:
                step4.mark_skip("テキストボックスが見つからなかったためスキップ")
                print(f"  [スキップ] {step4.step_name}")

            # --------------------------------------------------
            # Step 5: 送信ボタンを探す
            # --------------------------------------------------
            step5 = TestResult("送信ボタン検出")
            results.append(step5)
            submit_btn = None
            try:
                # 送信系ボタンを探す（優先順位付き）
                btn_selectors = [
                    'button:has-text("送信"):visible',
                    'button:has-text("保存"):visible',
                    'button:has-text("登録"):visible',
                    'button:has-text("追加"):visible',
                    'button:has-text("確認"):visible',
                    'input[type="submit"]:visible',
                    'button[type="submit"]:visible',
                ]
                for selector in btn_selectors:
                    locator = page.locator(selector).first
                    if locator.count() > 0 and locator.is_visible():
                        submit_btn = locator
                        break

                if submit_btn is None:
                    step5.mark_fail("送信ボタンが見つかりません")
                    step5.screenshot_path = take_screenshot(page, "04_no_button")
                    print(f"  [NG] {step5.step_name}")
                else:
                    btn_text = submit_btn.inner_text()
                    btn_id = submit_btn.evaluate("el => el.id || '(id無し)'")
                    step5.mark_pass(f"検出: '{btn_text}' id={btn_id}")
                    print(f"  [OK] {step5.step_name}: '{btn_text}'")
            except Exception as e:
                step5.mark_fail(f"ボタン検出エラー: {e}")
                step5.screenshot_path = take_screenshot(page, "04_button_error")
                print(f"  [NG] {step5.step_name}: {e}")

            # --------------------------------------------------
            # Step 6: 送信ボタンクリック
            # --------------------------------------------------
            step6 = TestResult("送信ボタンクリック")
            results.append(step6)
            if submit_btn is not None:
                try:
                    # クリック前のコンソールエラー数を記録
                    errors_before = len(console_errors)
                    submit_btn.click(timeout=ACTION_TIMEOUT)
                    # クリック後少し待つ（非同期処理を待つ）
                    page.wait_for_timeout(3000)
                    step6.mark_pass("ボタンクリック成功")
                    step6.screenshot_path = take_screenshot(page, "05_after_click")
                    print(f"  [OK] {step6.step_name}")
                except Exception as e:
                    step6.mark_fail(f"ボタンクリックエラー: {e}")
                    step6.screenshot_path = take_screenshot(page, "05_click_error")
                    print(f"  [NG] {step6.step_name}: {e}")
            else:
                step6.mark_skip("送信ボタンが見つからなかったためスキップ")
                print(f"  [スキップ] {step6.step_name}")

            # --------------------------------------------------
            # Step 7: エラー確認
            # --------------------------------------------------
            step7 = TestResult("エラー確認（画面上）")
            results.append(step7)
            try:
                # 画面上のエラー表示を確認
                error_selectors = [
                    '.error:visible',
                    '.alert-danger:visible',
                    '.error-message:visible',
                    '[role="alert"]:visible',
                    '.toast-error:visible',
                    '.swal2-popup:visible',  # SweetAlert2のポップアップ
                ]
                found_errors = []
                for selector in error_selectors:
                    elements = page.locator(selector)
                    count = elements.count()
                    for j in range(count):
                        el = elements.nth(j)
                        if el.is_visible():
                            text = el.inner_text().strip()
                            if text:
                                found_errors.append(f"[{selector}] {text[:100]}")

                if found_errors:
                    step7.mark_fail("画面上にエラー表示あり:\n" + "\n".join(found_errors))
                    step7.screenshot_path = take_screenshot(page, "06_ui_errors")
                    print(f"  [NG] {step7.step_name}: {len(found_errors)}件のエラー")
                else:
                    step7.mark_pass("画面上にエラー表示なし")
                    print(f"  [OK] {step7.step_name}")
            except Exception as e:
                step7.mark_fail(f"エラー確認中に例外: {e}")
                step7.screenshot_path = take_screenshot(page, "06_check_error")
                print(f"  [NG] {step7.step_name}: {e}")

            # --------------------------------------------------
            # Step 8: コンソールエラー確認
            # --------------------------------------------------
            step8 = TestResult("エラー確認（コンソール）")
            results.append(step8)
            if console_errors:
                step8.mark_fail(
                    f"コンソールエラー {len(console_errors)}件:\n"
                    + "\n".join(console_errors[:10])
                    + ("\n..." if len(console_errors) > 10 else "")
                )
                print(f"  [NG] {step8.step_name}: {len(console_errors)}件")
            else:
                step8.mark_pass("コンソールエラーなし")
                print(f"  [OK] {step8.step_name}")

            # --------------------------------------------------
            # 最終スクリーンショット
            # --------------------------------------------------
            take_screenshot(page, "99_final_state")

        finally:
            browser.close()

    # --------------------------------------------------
    # Excel出力
    # --------------------------------------------------
    write_results_to_excel(results, EXCEL_OUTPUT)

    # サマリー表示
    ok_count = sum(1 for r in results if r.status == "OK")
    ng_count = sum(1 for r in results if r.status == "NG")
    print("\n" + "=" * 60)
    print(f"テスト完了: OK={ok_count}, NG={ng_count}, 合計={len(results)}")
    print("=" * 60)

    return 1 if ng_count > 0 else 0


# ============================================================
# エントリーポイント
# ============================================================
if __name__ == "__main__":
    sys.exit(run_tests())
